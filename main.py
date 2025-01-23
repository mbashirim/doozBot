from telethon import TelegramClient, events, sync, Button
import json
import asyncio
import random
import threading
from datetime import datetime, timedelta
from game import FourInRow
from db import Database
from tron import TronManager
import json
import os
import time
import requests
from telethon.tl.types import KeyboardButton, ReplyKeyboardMarkup
from openpyxl import Workbook

with open('config.json', 'r') as file:
    config = json.load(file)
print(config)
api_id = config['auth_bot']['api_id']
api_hash = config['auth_bot']['api_hash']
bot_token = config['auth_bot']['token']
log_channel_id = config['log_channel_id']
db = Database()
client = TelegramClient('bot_session', api_id, api_hash)



def cleanup_old_games():
    while True:
        try:
            old_games = db.get_old_games()
            print(old_games)
            incomplete_games = db.get_incomplete_games(bot_token)
            if not old_games:
                time.sleep(10)
                continue
            print(f"Found {len(old_games)} old games to clean up")
            for game in old_games:
                game_id = game[0] 
                game_info = db.get_game_info_by_id(game_id)
                
                if not game_info:
                    continue
                    
                if game_info['current_player'] == '🔴':
                    print("Player 1 is the winner")
                    winner_id = game_info['player2_id'] 
                    loser_id = game_info['player1_id']
                else:
                    winner_id = game_info['player1_id']
                    loser_id = game_info['player2_id']
                    print("Player 2 is the winner")
                
                db.update_stats(str(winner_id), str(loser_id))
                
                winner_data = db.get_user(winner_id)
                loser_data = db.get_user(loser_id)
                
                winner_username = winner_data['username']
                loser_username = loser_data['username']
                if game_info['type_game'] == 'bet':
                    winner_amount = int(game_info['bet_amount']) - (game_info['bet_amount'] * 0.10) 
                    status = f"🏆 بازی رقابتی به دلیل بازی نکردن بازیکن {loser_username} به نفع بازیکن {winner_username} تمام شد و بازیکن {winner_username} برنده {winner_amount} سکه شد 🎉"
                else:
                    winner_amount = game_info['bet_amount']  
                    status = f"🏆 بازی دوستانه به دلیل بازی نکردن بازیکن {loser_username} به نفع بازیکن {winner_username} تمام شد و بازیکن {winner_username} برنده {winner_amount} سکه شد 🎉"
                
                if game_info['type_game'] == 'bet':
                    total_bet = int(game_info['bet_amount']) * 2
                    success, msg = db.add_coins_balance(winner_id, int(total_bet) - (int(total_bet) * 0.05))
                else:
                    total_bet = int(game_info['bet_amount']) * 2
                    success, msg = db.add_coins_balance_fr(winner_id, total_bet)

                for player_id, message_id in [(game_info['player1_id'], game_info['player1_message_id']),(game_info['player2_id'], game_info['player2_message_id'])]:
                    if not player_id or not message_id:
                        continue
                    url = f"https://api.telegram.org/bot{bot_token}/editMessageText"
                    data = {
                        'chat_id': player_id,
                        'message_id': message_id,
                        'text': status
                    }
                    response = requests.post(url, data=data)
                    if not response.ok:
                        print(f"Error editing message for player {player_id}: {response.text}")
                        continue

                requests.post(f"https://api.telegram.org/bot{bot_token}/sendMessage", data={
                    'chat_id': log_channel_id,
                    'text': f"🎮 بازی {game_id} با موفقیت به پایان رسید و بازیکن {winner_username} برنده شد!"
                })
                print(game_id)
                db.make_game_win(game_id, winner_id)
                    
            
            time.sleep(10)
            
        except Exception as e:
            print(f"Error in cleanup task: {e}")
            time.sleep(5)


cleanup_thread = threading.Thread(target=cleanup_old_games, daemon=True)
cleanup_thread.start()


async def check_channel_membership(event):
    """Check if user is member of required channel"""
    try:
        user_id = event.sender_id
        channel_username = "jhjhjuhuhkh"
        
        try:
            participant = await client.get_participants(channel_username, filter=None)
            is_member = any(p.id == user_id for p in participant)
        except:
            is_member = False
            
        if not is_member:
            join_button = Button.url("🔗 عضویت در کانال", f"https://t.me/{channel_username}")
            
            buttons = [[join_button]]
            
            await event.respond(
                "❌ برای استفاده از ربات باید عضو کانال رسمی ما باشید!\n\n"
                "👈 پس از عضویت، روی /start مجدد کلیک کنید.",
                buttons=buttons
            )
            return False
            
        return True
        
    except Exception as e:
        print(f"Error checking channel membership: {e}")
        return False
    
async def check_user_in_channel(user_id):
    """Check if specific user is member of channel"""
    try:
        user_id = int(user_id)
        channel_username = "fourinrow"
        
        try:
            participant = await client.get_participants(channel_username, filter=None)
            is_member = any(p.id == user_id for p in participant)
        except:
            is_member = False
            
        return is_member
        
    except Exception as e:
        print(f"Error checking channel membership: {e}")
        return False



games = {}
inline_games = {}
bet_amounts = {}  
pending_games = {} 
searching_players = {} 
search_timeouts = {} 
message_ids = {} 

@client.on(events.NewMessage(pattern='/start'))
async def start_handler(event):
    user_id = str(event.sender_id)
    username = event.sender.username
    
    user_data = db.get_user(user_id)
    if not user_data:
        message_text = event.message.text
        if len(message_text.split()) > 1:
            referrer_id = message_text.split()[1]
            if referrer_id != user_id:  
                db.add_coins_balance_fr(referrer_id, 1)
                db.save_user(user_id, username)
                try:
                    referral_count_file = 'referral_counts.json'
                    referral_lock = asyncio.Lock()

                    async with referral_lock:
                        if not os.path.exists(referral_count_file):
                            with open(referral_count_file, 'w') as f:
                                json.dump({}, f)

                        with open(referral_count_file, 'r') as f:
                            referral_counts = json.load(f)

                        if referrer_id in referral_counts:
                            referral_counts[referrer_id] += 1
                        else:
                            referral_counts[referrer_id] = 1

                        with open(referral_count_file, 'w') as f:
                            json.dump(referral_counts, f)
                    await client.send_message(int(referrer_id), "🎁 یک کاربر با لینک دعوت شما عضو شد!\n💰 1 سکه دوستانه به موجودی شما اضافه شد.")
                except:
                    pass
            else:
                db.save_user(user_id, username)
        else:
            db.save_user(user_id, username)
    
    if not user_data or not user_data.get('wallet_address'):
        db.create_wallet(user_id)
        
    if not await check_channel_membership(event):
        return

    keyboard = {
        "keyboard": [
            [{"text": "جستجوی حریف🕵‍♂"}, {"text": "دعوت حریف🥷"}],
            [{"text": "اطلاعات من👤"}, {"text": "معرفی👥"}], 
            [{"text": "اموزشی📚"}, {"text": "ارسال تیکت📨"}]
        ],
        "resize_keyboard": True
    }

    first_time_file = 'first_time_users.json'
    if not os.path.exists(first_time_file):
        with open(first_time_file, 'w') as f:
            json.dump({}, f)

    with open(first_time_file, 'r') as f:
        first_time_users = json.load(f)

    if user_id not in first_time_users:
        first_time_users[user_id] = str(datetime.now())
        with open(first_time_file, 'w') as f:
            json.dump(first_time_users, f)
            
        requests.post(
            f'https://api.telegram.org/bot{bot_token}/sendMessage',
            json={
                'chat_id': event.chat_id,
                'text': "به بازی دوز خوش آمدید! 🎮\n5 سکه به عنوان هدیه به شما داده شد.",
                'reply_markup': keyboard
            }
        )
    else:
        requests.post(
            f'https://api.telegram.org/bot{bot_token}/sendMessage',
            json={
                'chat_id': event.chat_id,
                'text': "به بازی دوز خوش آمدید! 🎮",
                'reply_markup': keyboard
            }
        )

async def search_timeout(user_id):
    await asyncio.sleep(60)
    if user_id in searching_players:
        del searching_players[user_id]
        if user_id in search_timeouts:
            del search_timeouts[user_id]
        try:
            await client.send_message(int(user_id), "تایم جستجو برای حریف به پایان رسید💣")
        except:
            pass

@client.on(events.NewMessage(pattern=r'^(/search|جستجوی حریف🕵‍♂)$'))
async def search_handler(event):
    user_id = str(event.sender_id)
    username = event.sender.username
    db = Database()
    active_games = db.get_active_games_json(user_id)
    print(active_games)
    invitions = db.get_invitions_of_player(user_id)
    print(invitions)
    if invitions >= 1:
        await event.respond("❌ شما نمی توانید بیشتر از 1 دعوت همزمان داشته باشید!")
        return
    total_active_games = active_games['bet'] + active_games['fr']
    if total_active_games >= 1:
        await event.respond("❌ شما نمی‌توانید بیش از 1 بازی همزمان داشته باشید!")
        return
    if not await check_channel_membership(event):
        return
    if user_id in searching_players:
        await event.respond("شما در حال حاضر در جستجوی حریف هستید❗️ (1 دقیقه باقی مانده)")
        return

    buttons = [
        [Button.inline("بازی دوستانه 🤝", f"search_friendly_{user_id}")],
        [Button.inline("بازی رقابتی 💰", f"search_bet_{user_id}")]
    ]
    
    await event.respond("🎮 شروع بازی 🎮", buttons=buttons)

@client.on(events.CallbackQuery(pattern=r"search_(friendly|bet)_.*"))
async def search_type_handler(event):
    data = event.data.decode().split('_')
    type_game = data[1]
    user_id = str(event.sender_id)
    
    searching_players[user_id] = {
        'type': type_game,
        'timestamp': datetime.now(),
        'username': event.sender.username
    }
    
    timeout_task = asyncio.create_task(search_timeout(user_id))
    search_timeouts[user_id] = timeout_task
    
    for search_id, search_data in list(searching_players.items()):
        if search_id != user_id and search_data['type'] == type_game:
            del searching_players[user_id]
            del searching_players[search_id]
            if user_id in search_timeouts:
                search_timeouts[user_id].cancel()
                del search_timeouts[user_id]
            if search_id in search_timeouts:
                search_timeouts[search_id].cancel() 
                del search_timeouts[search_id]
                
            game_type_text = "دوستانه" if type_game == "friendly" else "رقابتی"
            agreement_msg = f"""🎮 یک بازیکن پیدا شد!

👥 بازیکنان:
• @{event.sender.username}
• @{search_data['username']}

🎲 نوع بازی: {game_type_text}

روی مبلغ رقابتی توافق کنید."""
            bet_amount = 10 
            game_type_text = "رقابتی" if type_game == "bet" else "دوستانه"
            await event.edit(f"""حریف پیدا شد🥷

👥 بازیکنان:
• @{event.sender.username}
• @{search_data['username']}

نوع بازی: {game_type_text}🎲

منتظر پیشنهاد سکه حریف باشید💰
اگه شما پیشنهادی دارید در پیوی حریف به توافق برسید🤝""")

            try:
                await client.send_message(int(search_id), f"""🎮 بازی {game_type_text}
💰 پیشنهاد به حریف : {bet_amount} سکه

⚙️ برای تنظیم پیشنهاد خود از دکمه های زیر استفاده کنید👇""", buttons=[
                [Button.inline("-10", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{event.sender.id}_{bet_amount}_-10"),  Button.inline("-1", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{event.sender.id}_{bet_amount}_-1")],
                [Button.inline("+1", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{event.sender.id}_{bet_amount}_1"),  Button.inline("+10", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{event.sender.id}_{bet_amount}_10")],
                [Button.inline("تایید و ارسال دعوت ✅", f"confirm_{'fr' if type_game == 'friendly' else 'bet'}_{event.sender.id}_{bet_amount}")]
            ])
            except:
                pass
            return
            
    game_type = "دوستانه" if type_game == "friendly" else "رقابتی"
    await event.edit(f"در حال جستجوی حریف برای بازی {game_type}⏳")

@client.on(events.NewMessage(pattern=r'^(/invite|دعوت حریف🥷)$'))
async def invite_handler(event):
    user_id = str(event.sender_id)
    username = event.sender.username
    
    
    
    db = Database()
    invitions = db.get_invitions_of_player(user_id)
    print(invitions)
    if invitions >= 1:
        await event.respond("❌ شما نمی توانید بیشتر از 1 دعوت همزمان داشته باشید!")
        return
    active_games = db.get_active_games_json(user_id)
    total_active_games = active_games['bet'] + active_games['fr']
    
    if total_active_games >= 1:
        await event.respond("❌ شما نمی‌توانید بیش از 1 بازی همزمان داشته باشید!")
        return
    if not await check_channel_membership(event):
        return
    buttons = [
        [Button.inline("بازی دوستانه 🤝", f"invite_friendly_{user_id}")],
        [Button.inline("بازی رقابتی 💰", f"invite_bet_{user_id}")]
    ]
    
    await event.respond("🎮 شروع بازی 🎮", buttons=buttons)


@client.on(events.CallbackQuery(pattern=r"invite_(friendly|bet)_.*"))
async def invite_type_handler(event):
    data = event.data.decode().split('_')
    type_game = data[1]
    user_id = str(event.sender_id)
    await event.delete()
    
    db = Database()
    active_games = db.get_active_games_json(user_id)
    total_active_games = active_games['bet'] + active_games['fr']
    invitions = db.get_invitions_of_player(user_id)
    if invitions >= 1:
        await event.respond("❌ شما نمی توانید بیشتر از 1 دعوت همزمان داشته باشید!")
        return
    if total_active_games >= 1:
        await event.respond("❌ شما نمی‌توانید بیش از 1 بازی همزمان داشته باشید!")
        return
    await event.respond("آیدی حریفت رو بدون @ وارد کن👇")

    @client.on(events.NewMessage(from_users=[int(user_id)]))
    async def username_handler(event):
        invited_username = event.message.message.strip()
        invited_username_lower = invited_username.lower()
        invited_user_data = db.get_user_by_username(invited_username) or db.get_user_by_username(invited_username_lower)
        if not invited_user_data:
            await event.respond("کاربر مورد نظر یافت نشد❗️")
            client.remove_event_handler(username_handler)
            return
        if not await check_user_in_channel(invited_user_data[0]):
            await event.respond("❌ کاربر مورد نظر در کانال عضو نیست!")
            client.remove_event_handler(username_handler)
            return
        
        active_games_invited = db.get_active_games_json(invited_user_data[0])
        total_active_games_invited = active_games_invited['bet'] + active_games_invited['fr']
        if total_active_games_invited >= 1:
            await event.respond("❌ کاربر مورد نظر در حال حاظر یک بازی فعال دارد!")
            client.remove_event_handler(username_handler)
            return
        if not invited_user_data:
            await event.respond("کاربر مورد نظر یافت نشد❗️")
            client.remove_event_handler(username_handler)
            return
        if int(invited_user_data[0]) == int(event.sender_id):
            await event.respond("❌ شما نمی‌توانید خودتان را دعوت کنید!")
            client.remove_event_handler(username_handler)
            return
        invited_user_id = invited_user_data[0]
        
        bet_amount = 10 
        game_type_text = "رقابتی" if type_game == "bet" else "دوستانه"
        await event.respond(f"🎮 بازی {game_type_text}\n💰 سکه پیشنهادی: {bet_amount} سکه\nبا چند سکه میخای حریفت رو به بازی دعوت کنی؟😈", buttons=[
            [Button.inline("-10", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{invited_user_id}_{bet_amount}_-10"),  Button.inline("-1", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{invited_user_id}_{bet_amount}_-1")],
            [Button.inline("+1", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{invited_user_id}_{bet_amount}_1"),  Button.inline("+10", f"adjust_{'fr' if type_game == 'friendly' else 'bet'}_{invited_user_id}_{bet_amount}_10")],
            [Button.inline("تایید و ارسال دعوت ✅", f"confirm_{'fr' if type_game == 'friendly' else 'bet'}_{invited_user_id}_{bet_amount}")]
        ])
        client.remove_event_handler(username_handler)


@client.on(events.CallbackQuery(pattern=r"adjust_(bet|fr)_.*"))
async def adjust_bet_handler(event):
    data = event.data.decode().split('_')
    type_game = data[1]  
    invited_user_id = data[2]
    current_bet = int(data[3])
    adjustment = int(data[4])
    db = Database()
    active_games = db.get_active_games_json(invited_user_id)
    total_active_games = active_games['bet'] + active_games['fr']
    active_games2 = db.get_active_games_json(event.sender_id)
    total_active_games2 = active_games2['bet'] + active_games2['fr']
    invitions = db.get_invitions_of_player(event.sender_id)
    
    print(invitions)
    if invitions >= 1:
        await event.answer("❌ شما نمی توانید بیشتر از 1 دعوت همزمان داشته باشید!", alert=True)
        return
    
    if total_active_games2 >= 1:
        await event.answer("❌ شما نمی‌توانید بیش از 1 بازی همزمان داشته باشید!", alert=True)
        return
    if total_active_games >= 1:
        await event.answer("❌ کاربر مورد نظر در حال حاظر یک بازی فعال دارد!", alert=True)
        return
    new_bet = max(1, current_bet + adjustment)   
    game_type_text = "رقابتی" if type_game == "bet" else "دوستانه"
    await event.edit(f"🎮 بازی {game_type_text}\n💰 سکه پیشنهادی: {new_bet} سکه\nبا چند سکه میخای حریفت رو به بازی دعوت کنی؟😈", buttons=[
        [Button.inline("-10", f"adjust_{type_game}_{invited_user_id}_{new_bet}_-10"), Button.inline("-1", f"adjust_{type_game}_{invited_user_id}_{new_bet}_-1")],
        [Button.inline("+1", f"adjust_{type_game}_{invited_user_id}_{new_bet}_1"), Button.inline("+10", f"adjust_{type_game}_{invited_user_id}_{new_bet}_10")],
        [Button.inline("تایید و ارسال دعوت ✅", f"confirm_{type_game}_{invited_user_id}_{new_bet}")]
    ])

@client.on(events.CallbackQuery(pattern=r"confirm_(bet|fr)_.*"))
async def confirm_bet_handler(event):
    data = event.data.decode().split('_')
    type_game = data[1] 
    invited_user_id = data[2]
    bet_amount = int(data[3])
    user_id = str(event.sender_id)
    db = Database()
    active_games = db.get_active_games_json(user_id)
    total_active_games = active_games['bet'] + active_games['fr']
    invitions = db.get_invitions_of_player(user_id)
    
    if not await check_user_in_channel(user_id) or not await check_user_in_channel(invited_user_id):
        await event.answer("❌ شما یا کاربر دعوت شده در کانال عضو نیستید!", alert=True)
        return
    
    print(invitions)
    if invitions >= 1:
        await event.answer("❌ شما نمی توانید بیشتر از 1 دعوت همزمان داشته باشید!", alert=True)
        return
    if total_active_games >= 1:
        await event.answer("❌ شما نمی‌توانید بیش از 1 بازی همزمان داشته باشید!", alert=True)
        return
    user_data = db.get_user(user_id)
    balance_field = 'balance' if type_game == 'bet' else 'balance_fr'
    if not user_data or user_data[balance_field] < bet_amount:
        await event.answer("❌ موجودی شما برای این بازی کافی نیست!", alert=True)
        return
    invited_user_data = db.get_user(invited_user_id)
    invited_balance_field = 'balance' if type_game == 'bet' else 'balance_fr'
    if not invited_user_data or invited_user_data[invited_balance_field] < bet_amount:
        await event.answer("❌ موجودی کاربر دعوت شده برای این بازی کافی نیست!", alert=True)
        return
    
    
    game_id = str(random.randint(1000000, 9999999))
    game = FourInRow(player1_id=user_id)
    game.is_bet = True
    game.bet_amount = bet_amount
    game.type_game = type_game  
    inline_games[game_id] = game
    pending_games[game_id] = True
    game.save_to_db(game_id)
    
    type_game_text = "رقابتی" if type_game == 'bet' else "دوستانه"
    invite_message = (
        f"🎮 دعوت به بازی دوز {type_game_text}\n"
        f"💰پیشنهاد حریف: {bet_amount} سکه\n"
        f"👤 دعوت کننده: {user_data['username']}\n\n"
        "اگه شما پیشنهادی دارید در پیوی حریف به توافق برسید🤝"
    )
    buttons = [
        [Button.inline("شروع بازی🔥", f"accept_invite_{game_id}_{user_id}_{type_game}"), Button.inline("رد بازی❌", f"reject_invite_{game_id}_{user_id}_{type_game}")]
    ]
    await client.send_message(int(invited_user_id), invite_message, buttons=buttons)
    await event.edit("دعوت به حریف ارسال شد✅")

@client.on(events.CallbackQuery(pattern=r"accept_invite_.*"))
async def accept_invite_handler(event):
    db = Database()
    data = event.data.decode().split('_')
    print(data)
    game_id = data[2]
    inviter_id = data[3]
    type_game = data[4] 
    user_id = str(event.sender_id)
    user_data = db.get_user(user_id)

        
    game_info = db.get_game_by_game_id(game_id)
    if not game_info:
        await event.answer("دعوت ۳ دقیقه بعد ارسال منقضی شده است❗️", alert=True)
        await event.delete()
        return
        
    print(game_info)
    bet_amount = game_info[11]
    
    if type_game == 'bet':
        if user_data['balance'] < bet_amount:
            await event.answer("❌ موجودی کافی برای پیوستن به بازی رقابتی ندارید!", alert=True)
            await event.delete()
            await client.send_message(
                int(inviter_id),
                f"❌ بازیکن {user_data['username']} موجودی کافی برای پیوستن به بازی رقابتی را ندارد!"
            )
            db.delete_game_by_id(game_id)
            return
    else:  
        if user_data['balance_fr'] < bet_amount:
            await event.answer("❌ موجودی کافی برای پیوستن به بازی دوستانه ندارید!", alert=True)
            await event.delete()
            await client.send_message(
                int(inviter_id),
                f"❌ بازیکن {user_data['username']} موجودی کافی برای پیوستن به بازی دوستانه را ندارد!"
            )
            db.delete_game_by_id(game_id)
            return
    player1_games = db.get_active_games_count(inviter_id)
    player2_games = db.get_active_games_count(user_id)
    if player1_games >= 1:
        await event.answer("❌ بازیکن اول بیش از 1 بازی فعال دارد و نمی‌تواند بازی جدید ایجاد کند!", alert=True)
        await event.delete()
        db.delete_game_by_id(game_id)
        return
    if player2_games >= 1:
        await event.answer("❌ شما بیش از 1 بازی فعال دارید و نمی‌توانید به بازی جدید بپیوندید!", alert=True)
        await event.delete()
        db.delete_game_by_id(game_id)
        return
    if game_id not in inline_games:
        await event.answer("دعوت ۳ دقیقه بعد ارسال منقضی شده است❗️", alert=True)
        await event.delete()
        db.delete_game_by_id(game_id)
        return
    game = inline_games[game_id]
    if not db.join_player2(game_id, user_id):
        await event.answer("دعوت ۳ دقیقه بعد ارسال منقضی شده است❗️", alert=True)
        await event.delete()
        db.delete_game_by_id(game_id)
        return
    if game.is_bet:
        player2_data = db.get_user(user_id)
        if not player2_data:
            await event.answer("❌ خطا در دریافت اطلاعات کاربری!", alert=True)
            await event.delete()
            db.delete_game_by_id(game_id)
            return
        if type_game == 'bet':
            if player2_data['balance'] < game.bet_amount:
                await event.answer("❌ موجودی کافی برای پیوستن به بازی رقابتی ندارید!", alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
        else:
            if player2_data['balance_fr'] < game.bet_amount:
                await event.answer("❌ موجودی کافی برای پیوستن به بازی دوستانه ندارید!", alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
        player1_data = db.get_user(inviter_id)
        if not player1_data:
            await event.answer("❌ خطا در دریافت اطلاعات بازیکن اول!", alert=True)
            await event.delete()
            db.delete_game_by_id(game_id)
            return
        if type_game == 'bet':
            if player1_data['balance'] < game.bet_amount:
                await event.answer(f"❌ @{player1_data['username']} موجودی کافی برای پیوستن به بازی را ندارد❗️", alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
        else:
            if player1_data['balance_fr'] < game.bet_amount:
                await event.answer(f"❌ @{player1_data['username']} موجودی کافی برای پیوستن به بازی را ندارد❗️", alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
        if type_game == 'bet':
            success1, msg1 = db.remove_coins(inviter_id, game.bet_amount)
            if not success1:
                await event.answer("❌ خطا در کسر سکه از بازیکن اول: " + msg1, alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
                
            success2, msg2 = db.remove_coins(user_id, game.bet_amount)
            if not success2:
                db.add_coins_balance(inviter_id, game.bet_amount)
                await event.answer("❌ خطا در کسر سکه از بازیکن دوم: " + msg2, alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
        else:
            success1, msg1 = db.remove_coins_fr(inviter_id, game.bet_amount)
            if not success1:
                await event.answer("❌ خطا در کسر سکه از بازیکن اول: " + msg1, alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return
                
            success2, msg2 = db.remove_coins_fr(user_id, game.bet_amount)
            if not success2:
                db.add_coins_balance_fr(inviter_id, game.bet_amount)
                await event.answer("❌ خطا در کسر سکه از بازیکن دوم: " + msg2, alert=True)
                await event.delete()
                db.delete_game_by_id(game_id)
                return

    game.player2_id = user_id
    game.player2_username = event.sender.username
    username = event.sender.username
    db.save_user(user_id, username)
    
    if game_id in pending_games:
        del pending_games[game_id]
        
        await event.delete()
        
        bet_type = "رقابتی" if game.type_game == "bet" else "دوستانه"
        game_message = f"""بازیکن دوم پیوست! بازی شروع شد!
💰 مبلغ بازی: {game.bet_amount} سکه
🎮 نوع بازی: {bet_type}
نوبت: {game.current_player} {game.player2_username if game.current_player == '🔵' else game.player1_username}"""
        try:
            msg1 = await client.send_message(int(game.player1_id), game_message, buttons=game.get_board_buttons(game_id))
            msg2 = await client.send_message(int(game.player2_id), game_message, buttons=game.get_board_buttons(game_id))
            
            print(game_id, game.player1_id, msg1.id)
            db.save_message_id_for_user(game_id, game.player1_id, msg1.id)
            db.save_message_id_for_user(game_id, game.player2_id, msg2.id)
            message_ids[game_id] = {
                game.player1_id: msg1.id,
                game.player2_id: msg2.id
            }
            await client.send_message(log_channel_id, f"🎮 بازی {game_id} شروع شد!\n👤 بازیکن اول: {game.player1_username}\n👤 بازیکن دوم: {game.player2_username}\n💰 مبلغ شرط: {game.bet_amount} سکه\n🎮 نوبت: {game.current_player}")
        except Exception as e:
            print(f"Error notifying players: {e}")

@client.on(events.CallbackQuery(pattern=r"reject_invite_.*"))
async def reject_invite_handler(event):
    await event.delete()
    data = event.data.decode().split('_')
    game_id = data[2]
    inviter_id = data[3]
    
    if game_id in inline_games:
        game = inline_games[game_id]
        del inline_games[game_id]

    db = Database()
    success, msg = db.delete_game_by_id(game_id)
    if not success:
        print(f"Error deleting game {game_id}: {msg}")
        
    if game_id in pending_games:
        del pending_games[game_id]
    await event.answer("دعوت‌نامه رد شد!", alert=True)
    await client.send_message(int(inviter_id), "دعوت شما رد شد❗️")
    await client.send_message(log_channel_id, f"❌ دعوت‌نامه بازی {game_id} توسط کاربر {event.sender_id} رد شد.")


@client.on(events.CallbackQuery(pattern=r"join_.*"))
async def join_game_handler(event):
    user_id = str(event.sender_id)
    game_id = event.data.decode().split('_')[1]
    
    if game_id not in inline_games:
        await event.answer("این بازی دیگر در دسترس نیست!", alert=True)
        return
        
    game = inline_games[game_id]
    
    if user_id == game.player1_id:
        await event.answer("شما نمیتوانید به بازی خودتان بپیوندید!", alert=True)
        return
        
    if game.is_bet:
        player2_data = db.get_user(user_id)
        if not player2_data or player2_data['balance'] < game.bet_amount:
            await event.answer("❌ موجودی کافی برای پیوستن به بازی رقابتی ندارید!", alert=True)
            return
        
    game.player2_id = user_id
    game.player2_username = event.sender.username
    username = event.sender.username
    db.save_user(user_id, username)
    
    game.save_to_db(game_id)
    
    if game_id in pending_games:
        del pending_games[game_id]
        game_mode_text = "رقابتی" if game.is_bet else "دوستانه"
        msg = await event.edit(
            f"حریف پیوست😈\n"
            f"بازی شروع شد🎲\n"
            f"{'💰' if game.is_bet else '🔮'}تعداد سکه: {game.bet_amount} سکه\n"
            f"🎮بازی: {game_mode_text}\n"
            f"⚔️نوبت: {game.current_player} {game.player1_username if game.current_player == '🔴' else game.player2_username}",
            buttons=game.get_board_buttons(game_id)
        )
        
        message_ids[game_id] = {
            game.player1_id: msg.id,
            game.player2_id: msg.id
        }

@client.on(events.CallbackQuery())
async def button_handler(event):
    user_id = str(event.sender_id)
    username = event.sender.username
    db = Database()
    db.save_user(user_id, username)

    data = event.data.decode().split('_')
    
    if data[0] == 'status':
        await event.answer("بازی تمام شده است!")
        return
    
    if len(data) == 3:
        game_id, row, col = data
        game_id = str(game_id)
        
        if row == 'top':
            col = int(col)
        else:
            row = int(row)
            col = int(col)
            
        if game_id not in inline_games:
            try:
                game = FourInRow.load_from_db(game_id)
                if not game:
                    await event.answer("این بازی منقضی شده است!")
                    return
                inline_games[game_id] = game
            except ValueError:
                await event.answer("خطا: شناسه بازی نامعتبر است!")
                return
            except Exception as e:
                await event.answer(f"خطای غیرمنتظره: {str(e)}")
                return
            
        game = inline_games[game_id]
        
        if game_id in pending_games:
            await event.answer("لطفا منتظر پیوستن بازیکن دوم باشید!", alert=True)
            return
            
        current_player_id = game.player1_id if game.current_player == '🔴' else game.player2_id
        if str(event.sender_id) != current_player_id:
            await event.answer("نوبت شما نیست!")
            return
            
        if not game.make_move(col):
            await event.answer("حرکت نامعتبر! دوباره تلاش کنید.")
            return

        game.save_to_db(game_id)

        if game.game_over:
            if game.moves == 42:
                message = "❗️بازی مساوی شد❗️"
            else:
                winner_id = game.player1_id if game.current_player == '🔵' else game.player2_id
                loser_id = game.player2_id if game.current_player == '🔵' else game.player1_id
                winner_name = game.player1_username if winner_id != game.player1_id else game.player2_username
                message = f"🏆 بازیکن {winner_name} برنده بازی شد! 🎉"
                
                if game.type_game == 'bet':
                    total_bet = (game.bet_amount) - ((game.bet_amount) * 0.10)
                    message += f"\n💰 برنده {total_bet} سکه دریافت کرد!"
                else:
                    total_bet = game.bet_amount
                    message += f"\n💰 برنده {total_bet} سکه دریافت کرد!"
                    
                db = Database()
                results = db.get_games_between_players(game.player1_id, game.player2_id)
                if game.type_game == 'fr':
                    message += "\n\n⚔️ نتایج کل مسابقات دوستانه بین شما دو نفر:\n"
                    message += f"{game.player1_username}: {results['fr'][game.player1_id]}\n"
                    message += f"{game.player2_username}: {results['fr'][game.player2_id]}"
                elif game.type_game == 'bet':
                    message += "\n\n⚔️ نتایج کل مسابقات رقابتی بین شما دو نفر:\n"
                    message += f"{game.player1_username}: {results['win'][game.player1_id]}\n"
                    message += f"{game.player2_username}: {results['win'][game.player2_id]}"
                
                log_message = f"🎮 بازی {game_id} به پایان رسید!\n" \
f"👤 بازیکن اول: {game.player1_username}\n" \
f"👤 بازیکن دوم: {game.player2_username}\n" \
f"💰 مبلغ شرط: {game.bet_amount} سکه\n" \
f"🎮 حالت بازی: {'رقابتی' if game.type_game == 'bet' else 'دوستانه'}\n" \
f"🏆 برنده: {winner_name}\n"
                await client.send_message(log_channel_id, log_message)
                
            for player_id in [game.player1_id, game.player2_id]:
                try:
                    if game_id in message_ids and player_id in message_ids[game_id]:
                        await client.edit_message(
                            int(player_id),
                            message_ids[game_id][player_id],
                            message,
                            buttons=game.get_board_buttons(game_id)
                        )
                except Exception as e:
                    print(f"Error updating board for player {player_id}: {e}")
                    
            del inline_games[game_id]
            if game_id in message_ids:
                del message_ids[game_id]
            return
            
        for player_id in [game.player1_id, game.player2_id]:
            try:
                if game_id in message_ids and player_id in message_ids[game_id]:
                    await client.edit_message(
                        int(player_id),
                        message_ids[game_id][player_id],
                        f"⚔️نوبت: {game.current_player} {game.player2_username if game.current_player == '🔵' else game.player1_username}\n"
                        f"{'🔮' if game.type_game == 'fr' else '💰'}تعداد سکه: {game.bet_amount} سکه\n"
                        f"🎮بازی: {'رقابتی' if game.type_game == 'bet' else 'دوستانه'}",
                        buttons=game.get_board_buttons(game_id)
                    )
            except Exception as e:
                print(f"Error updating board for player {player_id}: {e}")

@client.on(events.CallbackQuery(pattern=r"play_again_.*"))
async def play_again_handler(event):
    data = event.data.decode().split('_')
    player1_id = data[2]
    player2_id = data[3]
    bet_amount = int(data[4])
    game_type = data[5]
    
    user_id = str(event.sender_id)
    
    if user_id not in [player1_id, player2_id]:
        await event.answer("شما مجاز به شروع این بازی نیستید!", alert=True)
        return
    
    opponent_id = player2_id if user_id == player1_id else player1_id
    game_type_text = "رقابتی" if game_type == "bet" else "دوستانه"
    
    await event.respond(f"🎮 بازی {game_type_text}\n💰 سکه پیشنهادی: {bet_amount} سکه\n\nبا چند سکه میخای حریفت رو به بازی دعوت کنی؟😈", buttons=[
        [Button.inline("-10", f"adjust_{game_type}_{opponent_id}_{bet_amount}_-10"), 
         Button.inline("-1", f"adjust_{game_type}_{opponent_id}_{bet_amount}_-1")],
        [Button.inline("+1", f"adjust_{game_type}_{opponent_id}_{bet_amount}_1"), 
         Button.inline("+10", f"adjust_{game_type}_{opponent_id}_{bet_amount}_10")],
        [Button.inline("تایید و ارسال دعوت ✅", f"confirm_{game_type}_{opponent_id}_{bet_amount}")]
    ])

@client.on(events.NewMessage(pattern=r'^(/info|اطلاعات من👤)$'))
async def info_command(event):
    try:
        user_id = str(event.sender_id)
        if not await check_channel_membership(event):
            return
        db = Database()
        active_games = db.get_active_games_json(user_id)
        total_active_games = active_games['bet'] + active_games['fr']
        
        if total_active_games >= 1:
            await event.respond("اطلاعات شما بعد از پایان بازی بروز رسانی میشه‼️")
            return
        user_data = db.get_user(user_id)
        if not user_data:
            await event.reply("❌ اطلاعات کاربری شما یافت نشد!")
            return

        game_info = db.get_full_game_info(user_id)
        if not game_info:
            await event.reply("❌ خطا در دریافت اطلاعات بازی!")
            return
        active_games = db.get_active_games_json(user_id)

        info_text = f"""**📊 اطلاعات کاربری شما:**

**👤 شناسه:** `{user_data['id']}`
**🎮 نام کاربری:** `{user_data['username'] or 'تنظیم نشده'}`
**💰 موجودی سکه رقابتی:** `{float(user_data.get('balance', 0)):.2f} 🪙`
**🔮 موجودی سکه دوستانه:** `{float(user_data.get('balance_fr', 0)):.2f} 🪙`

**📈 آمار بازی:**

**دوستانه:**
**🏆 برد:** `{game_info['fr']['win']}`
**💔 باخت:** `{game_info['fr']['lose']}`
**🤝 مساوی:** `{game_info['fr']['draw']}`

**رقابتی:**
**🏆 برد:** `{game_info['bet']['win']}`
**💔 باخت:** `{game_info['bet']['lose']}`
**🤝 مساوی:** `{game_info['bet']['draw']}`

**🎮 بازی های فعال:**
**رقابتی:** `{active_games['bet']}`
**دوستانه:** `{active_games['fr']}`"""



        charge_buttons = [
            [Button.inline("💰 شارژ سکه رقابتی", f"charge_{user_id}")],
            [Button.inline("🔮 شارژ مجدد سکه دوستانه", f"chargefr_{user_id}")],
            [Button.inline("💸 برداشت سکه رقابتی", f"withdraw_{user_id}")],
        ]
        
        await event.reply(info_text, buttons=charge_buttons, parse_mode='markdown')
        
    except Exception as e:
        print(f"Error in info command: {e}")
        await event.reply("❌ خطا در دریافت اطلاعات!")

@client.on(events.CallbackQuery(pattern=r'^withdraw_'))
async def withdraw_callback(event):
    try:
        user_id = event.data.decode().split('_')[1]
        user_data = db.get_user(user_id)
        wallet = ""
        if not user_data:
            await event.answer("❌ اطلاعات کاربری شما یافت نشد!", alert=True)
            return
            
        if user_data['balance'] < 1:
            await event.answer("❌ حداقل موجودی برای برداشت 1 سکه است!", alert=True)
            return

        await event.edit(
            "💳 برداشت موجودی به کیف پول TRX\n\n"
            "🔗 لطفا آدرس کیف پول TRX خود را وارد کنید:\n\n"
            "⚠️ نکات مهم:\n"
            "• تمام موجودی سکه شما برداشت خواهد شد\n" 
            "• مسئولیت اشتباه در وارد کردن آدرس کیف پول با خودتان است\n"
            "• فقط آدرس‌های معتبر TRX (شروع با T) قابل قبول است"
        )
        
        @client.on(events.NewMessage(from_users=[int(user_id)]))
        async def wallet_handler(event):
            wallet = event.message.message.strip()
            
            if not wallet.startswith('T') or len(wallet) != 34:
                await event.respond("❌ آدرس کیف پول نامعتبر است! لطفا یک آدرس معتبر TRX وارد کنید.")
                client.remove_event_handler(wallet_handler)
                return
                
            user_data = db.get_user(str(event.sender_id))
            if not user_data:
                await event.respond("❌ خطا در دریافت اطلاعات کاربری!")
                client.remove_event_handler(wallet_handler)
                return
                
            balance = user_data['balance']
            rate = 5
            trx_amount = (balance * rate) - 2.2
            print(trx_amount)
                
            try:
                tron = TronManager()
                trx_amount = round(float(trx_amount), 2)
                
                tron.send_trx(config['bank_wallet']['address'], config['bank_wallet']['private_key'], wallet, trx_amount)
                db.reset_coins(user_id)
                
                trx_equiv = round(balance * rate, 2)
                final_amount = round((balance * rate) - 2.2, 2)
                
                await client.send_message(log_channel_id, 
                    f"💰 برداشت موجودی کاربر {user_id}:\n\n"
                    f"💰 تعداد سکه: {balance}\n"
                    f"💎 معادل ترون: {trx_equiv:.2f}\n"
                    f"🔸 مبلغ نهایی پس از کسر کارمزد: {final_amount:.2f} TRX\n\n"
                    f"🏦 آدرس کیف پول: {wallet}")
                
                await event.respond(
                    f"✅ درخواست برداشت شما با موفقیت ثبت شد!\n\n"
                    f"💰 تعداد سکه: {balance}\n"
                    f"💸 معادل ترون: {trx_equiv:.2f}\n" 
                    f"💥 مبلغ نهایی پس از کسر کارمزد: {final_amount:.2f} TRX\n\n"
                    f"🏦 آدرس کیف پول: {wallet}\n\n"
                    f"⏳ برداشت شما در حال پردازش است و تا ساعاتی دیگر به کیف پولتان واریز میشود")
                
                client.remove_event_handler(wallet_handler)
                
            except (ValueError, TypeError) as e:
                print(f"Error converting values: {e}")
                await event.respond("❌ خطا در محاسبه مقادیر. لطفا با پشتیبانی تماس بگیرید.")
                client.remove_event_handler(wallet_handler)
        
        print(wallet)
        
        
        
        
    except Exception as e:
        print(f"Error in withdraw callback: {e}")
        await event.answer("❌ خطا در پردازش درخواست برداشت!", alert=True)





@client.on(events.CallbackQuery(pattern=r'^chargefr_'))
async def chargefr_callback(event):
    try:
        user_id = event.data.decode().split('_')[1]
        db = Database()
        active_games = db.get_active_games_json(user_id)
        total_active_games = active_games['bet'] + active_games['fr']
        
        if total_active_games >= 1:
            await event.answer("❌ شما بازی فعال دارید! بعد از اتمام بازی فعلی می‌توانید مجدداً شارژ کنید.", alert=True)
            return
        user_data = db.get_user(user_id)
        if user_data and user_data['balance_fr'] == 0:
            success, message = db.add_coins_balance_fr(user_id, 5)
            if success:
                await event.answer("موجودی دوستانه شما با موفقیت شارژ شد!", alert=True)
                await client.send_message(log_channel_id, f"💰 کاربر موجودی دوستانه شارژ کرد: {5} 🪙")
            else:
                await event.answer(f"خطا در شارژ موجودی دوستانه: {message}", alert=True)
        else:
            await event.answer("شما سکه دوستانه برای بازی دارید❗️", alert=True)
    except Exception as e:
        print(f"Error in chargefr callback: {e}")
        await event.answer("خطا در شارژ موجودی دوستانه!", alert=True)

@client.on(events.NewMessage(pattern=r'^(/learn|اموزشی📚)$'))
async def learn_command(event):
    if not await check_channel_membership(event):
        return
    try:
        tutorial_text = """🎮 **آموزش بازی دوز 4 تایی**

**هدف بازی:**
هدف بازی قرار دادن 4 مهره همرنگ به صورت افقی، عمودی یا مورب است.

**نحوه بازی:**
1️⃣ بازیکنان به نوبت مهره‌های خود را در ستون‌های بازی قرار می‌دهند
2️⃣ مهره‌ها از پایین ستون به بالا چیده می‌شوند
3️⃣ اولین بازیکنی که بتواند 4 مهره همرنگ را در یک راستا قرار دهد، برنده است!

**نکات مهم:**
• در هر نوبت فقط یک مهره می‌توانید قرار دهید
• به استراتژی حریف دقت کنید و مانع برد او شوید
• سعی کنید چند مسیر برد برای خود ایجاد کنید

**🤖 آموزش استفاده از ربات:**

**1️⃣ شروع بازی:**
• با دستور /start بازی را شروع کنید
• از منوی اصلی یکی از گزینه‌ها را انتخاب کنید

🤝 **بازی دوستانه:**
• با سکه‌های دوستانه بازی کنید
• محیطی مناسب برای تمرین و یادگیری
• فرصتی برای آشنایی با قلق‌های بازی

💰 **بازی رقابتی:**
• با سکه‌های واقعی به رقابت بپردازید
• محک زدن مهارت‌های خود در محیط رقابتی
• فرصت کسب جایزه

**5️⃣ نکات مهم:**
• قبل از ورود به بازی‌های رقابتی، در بخش دوستانه تمرین کنید
• با تمرین بیشتر، تاکتیک‌های پیشرفته را یاد بگیرید
• با هر چند سکه که با رقیبتون به توافق رسیدید میتوانید رقابت کنید"""

        await client.send_file(
            event.chat_id,
            "tutorial.mp4",
            caption=tutorial_text,
            parse_mode='markdown'
        )

    except Exception as e:
        print(f"Error in learn command: {e}")
        await event.reply("❌ خطا در نمایش آموزش!")

@client.on(events.CallbackQuery(pattern=r"charge_.*"))
async def charge_callback(event):
    try:
        user_id = event.data.decode().split('_')[1]
        wallet = db.get_wallet(user_id)
        
        if not wallet:
            await event.answer("❌ کیف پول شما یافت نشد!", alert=True)
            return
        
        await event.reply("چند سکه میخواهید شارژ کنید⁉️")

        @client.on(events.NewMessage(from_users=int(user_id)))
        async def handle_charge_amount(event):
            try:
                amount = int(event.message.message)
                if amount <= 0:
                    await event.reply("❗️حداقل واریزی ۱ سکه‌ معادل ۵ ترون است")
                    return
                
                trx_amount = amount * 5  
                
                wallet_info = f"""💳 اطلاعات شارژ کیف پول:

🏦 آدرس کیف پول: 
`{wallet}`

💰 نرخ تبدیل:
`{amount} سکه = {trx_amount} TRX`

⚠️ هشدار مهم:
• در صورت ارسال مقدار کمتر از {trx_amount} TRX، تراکنش انجام نخواهد شد
• در صورت ارسال مقدار بیشتر از {trx_amount} TRX، مابقی مبلغ برگشت داده نمی‌شود
• لطفا دقیقا مقدار {trx_amount} TRX را ارسال کنید

**برای شارژ کیف پول ربات،ترون را به آدرس بالا واریز کنید.**
**بعد از واریز،دکمه تایید تراکنش رو بزنید موجودی شما آپدیت خواهد شد.**"""

                charge_button = Button.inline("✅ تایید تراکنش ✅", f"check_{user_id}")
                
                await event.reply(wallet_info, buttons=charge_button, parse_mode='markdown')
                client.remove_event_handler(handle_charge_amount)   
            except ValueError:
                client.remove_event_handler(handle_charge_amount)   
                pass
            except Exception as e:
                client.remove_event_handler(handle_charge_amount)   
                print(f"Error in handle charge amount: {e}")
                await event.reply("❌ خطا در پردازش مقدار شارژ!") 
    except Exception as e:
        print(f"Error in charge callback: {e}")
        await event.answer("❌ خطا در دریافت اطلاعات کیف پول!", alert=True)

@client.on(events.CallbackQuery(pattern=r"check_.*"))
async def check_balance_callback(event):
    try:
        user_id = event.data.decode().split('_')[1]
        wallet = db.get_wallet(user_id)
        
        if not wallet:
            await event.answer("❌ کیف پول شما یافت نشد!", alert=True)
            return
            
        tron = TronManager()
        trx_balance = tron.get_trx_balance(wallet)
        
        if trx_balance > 1.1:
            coins = int((trx_balance) / 5)
            if db.add_coins(user_id, coins):
                success_msg = f"""✅ تراکنش شما با موفقیت انجام شد!

{coins} 🪙

موجودی شما به‌روز شد."""
                await event.edit(success_msg)
                await client.send_message(log_channel_id, f"💰 کاربر شارژ کرد: {coins} 🪙")
            else:
                await event.answer("❌ خطا در به‌روزرسانی موجودی!", alert=True)
        else:
            await event.answer("❌ تراکنشی یافت نشد! لطفا پس از واریز TRX دوباره تلاش کنید.", alert=True)
            
    except Exception as e:
        print(f"Error in check balance callback: {e}")
        await event.answer("❌ خطا در بررسی تراکنش!", alert=True)

@client.on(events.NewMessage(pattern=r"/bank_balance"))
async def bank_balance_callback(event):
    try:
        sender = await event.get_sender()
        if sender.username != 'mensurscars':
            await event.reply("❌ شما دسترسی به این بخش را ندارید!")
            return

        with open('config.json') as f:
            config = json.load(f)
        bank_address = config['bank_wallet']['address']

        tron = TronManager()
        bank_balance = tron.get_trx_balance(bank_address)

        balance_msg = f"""💰 موجودی کیف پول بانک:

{bank_balance} TRX

در صورتی که ولت دوم رو داخل فایل کانفیگ ربات ثبت کرده باشید و قصد نداشتین که مستقیم از ولت بانک استفاده کنین از دکمه زیر استفاده کنین"""

        claim_button = Button.inline("💰 انتقال به ولت دوم", "bank_claim_confirm")
        await event.reply(balance_msg, buttons=claim_button)

    except Exception as e:
        print(f"Error in bank balance callback: {e}")
        await event.reply("❌ خطا در دریافت موجودی بانک!")

@client.on(events.CallbackQuery(pattern=r"bank_claim_confirm"))
async def bank_claim_confirm_callback(event):
    try:
        sender = await event.get_sender()
        if sender.username != 'mensurscars':
            await event.answer("❌ شما دسترسی به این بخش را ندارید!", alert=True)
            return

        db = Database()
        all_games = db.get_all_bet_games()
        
        try:
            with open('claimed_games.json', 'r') as f:
                claimed_games = json.load(f)
        except FileNotFoundError:
            claimed_games = []
            with open('claimed_games.json', 'w') as f:
                json.dump(claimed_games, f)
                
        total_trx = 0
        unclaimed_games = []
        
        for game in all_games:
            game_id = game[0]
            bet_amount = game[11]
            
            if game_id not in claimed_games and bet_amount:
                price_in_trx = bet_amount * 5
                owner_cut = price_in_trx * 0.1  # 10% cut
                total_trx += owner_cut
                unclaimed_games.append(game_id)
                
        final_amount = total_trx -  1.1
        
        with open('config.json') as f:
            config = json.load(f)
        dest_address = config['destenation_wallet']['address']
        
        confirm_msg = f"""📊 اطلاعات انتقال موجودی:

💰 تعداد بازی‌های پرداخت نشده: {len(unclaimed_games)}
💎 کل TRX قابل برداشت: {total_trx:.2f}
🔸 کسر فی انتقال: {(total_trx) - 1.1:.2f} TRX
✨ مبلغ نهایی: {final_amount:.2f} TRX
🏦 آدرس مقصد: {dest_address}

آیا از انتقال موجودی به ولت دوم اطمینان دارید؟"""
        confirm_buttons = [
            [Button.inline("✅ بله، انتقال بده", "bank_claim_yes"),Button.inline("❌ خیر، لغو", "bank_claim_no")]
        ]
        await event.edit(confirm_msg, buttons=confirm_buttons)

    except Exception as e:
        print(f"Error in bank claim confirm callback: {e}")
        await event.answer("❌ خطا در تایید انتقال!", alert=True)

@client.on(events.CallbackQuery(pattern=r"bank_claim_no"))
async def cancel_claim_callback(event):
    try:
        sender = await event.get_sender()
        if sender.username != 'mensurscars':
            await event.answer("❌ شما دسترسی به این بخش را ندارید!", alert=True)
            return
            
        await event.edit("❌ انتقال موجودی لغو شد.")
        
    except Exception as e:
        print(f"Error in cancel claim callback: {e}")
        await event.answer("❌ خطا در لغو انتقال!", alert=True)

@client.on(events.CallbackQuery(pattern=r"bank_claim_yes"))
async def bank_claim_callback(event):
    try:
        sender = await event.get_sender()
    

        with open('config.json') as f:
            config = json.load(f)
        bank_address = config['bank_wallet']['address']
        bank_key = config['bank_wallet']['private_key']
        dest_address = config['destenation_wallet']['address']
        owner_username = config['owner_username']
        if sender.username != owner_username:
            await event.answer("❌ شما دسترسی به این بخش را ندارید!", alert=True)
            return


        db = Database()
        all_games = db.get_all_bet_games()
        
        if not os.path.exists('claimed_games.json'):
            claimed_games = []
            with open('claimed_games.json', 'w') as f:
                json.dump(claimed_games, f)
                
        with open('claimed_games.json', 'r') as f:
            claimed_games = json.load(f)

        tron = TronManager()
        bank_balance = tron.get_trx_balance(bank_address)
        
        if bank_balance > 1.1:
            try:
                with open('claimed_games.json', 'r') as f:
                    claimed_games = json.load(f)
            except FileNotFoundError:
                claimed_games = []
                with open('claimed_games.json', 'w') as f:
                    json.dump(claimed_games, f)
                    
            total_trx = 0
            unclaimed_games = []
            
            for game in all_games:
                game_id = game[0]
                bet_amount = game[11]
                
                if game_id not in claimed_games and bet_amount:
                    price_in_trx = bet_amount * 5
                    owner_cut = price_in_trx * 0.1  # 10% cut
                    total_trx += owner_cut
                    unclaimed_games.append(game_id)
                    
            transfer_amount = total_trx
            result = tron.send_trx(bank_address, bank_key, dest_address, transfer_amount)
            
            # Update claimed games
            for game in all_games:
                game_id = game[0]
                if game_id not in claimed_games:
                    claimed_games.append(game_id)
                    
            with open('claimed_games.json', 'w') as f:
                json.dump(claimed_games, f)
            
            success_msg = f"""✅ انتقال با موفقیت انجام شد!
            
💰 مقدار {transfer_amount} TRX به ولت دوم منتقل شد."""
            await event.edit(success_msg)
        else:
            await event.edit("❌ موجودی کافی نیست! حداقل 1.1 TRX برای کارمزد نیاز است.")

    except Exception as e:
        print(f"Error in bank claim callback: {e}")
        await event.edit("❌ خطا در انتقال به ولت دوم!")

@client.on(events.NewMessage(pattern=r"^(/referral|معرفی👥)$"))
async def referral_callback(event):
    if not await check_channel_membership(event):
        return
    try:
        sender = await event.get_sender()
        user_id = sender.id

        with open('config.json') as f:
            config = json.load(f)
            bot_username = config['bot_username']
        referral_link = f"https://t.me/{bot_username}?start={user_id}"

        referral_count_file = 'referral_counts.json'
        if not os.path.exists(referral_count_file):
            with open(referral_count_file, 'w') as f:
                json.dump({}, f)

        with open(referral_count_file, 'r') as f:
            referral_counts = json.load(f)

        referral_count = referral_counts.get(str(user_id), 0)

        msg = f"""👥 سیستم معرفی

🔗 لینک دعوت شما:
`{referral_link}`

💰 پاداش دعوت:
با دعوت هر دوست به ربات، `1` سکه دوستانه دریافت کنید!

📌 راهنما:
1. لینک دعوت خود را برای دوستانتان ارسال کنید
2. دوستان شما باید روی لینک کلیک کرده و ربات را استارت کنند
3.دوستان شما باید در کانال ربات عضو شوند

به ازای هر دوست جدید 1 سکه دوستانه به شما اضافه می‌شود🎁

هر چه دوستان بیشتری دعوت کنید، سکه‌های بیشتری دریافت می‌کنید🔥

📈 تعداد دعوت‌های شما: `{referral_count}`"""

        share_button = Button.url("🔗 اشتراک گذاری با دوستان", f"https://t.me/share/url?url={referral_link}")

        if isinstance(event, events.CallbackQuery.Event):
            await event.edit(msg, buttons=share_button, parse_mode='markdown')
        else:
            await event.reply(msg, buttons=share_button, parse_mode='markdown')

    except Exception as e:
        print(f"Error in referral callback: {e}")
        if isinstance(event, events.CallbackQuery.Event):
            await event.answer("❌ خطا در نمایش اطلاعات زیرمجموعه!", alert=True)
        else:
            await event.reply("❌ خطا در نمایش اطلاعات زیرمجموعه!")

@client.on(events.NewMessage(pattern=r'^(/ticket|ارسال تیکت📨)$'))
async def ticket_handler(event):
    if not await check_channel_membership(event):
        return

    try:
        async with client.conversation(event.chat_id, timeout=300) as conv:
            await conv.send_message("📝 لطفا پیام خود را ارسال کنید:")
            
            try:
                user_response = await conv.get_response()
                message = user_response.text.strip()
                
                sender = await event.get_sender()
                user_id = sender.id
                username = sender.username or "بدون نام کاربری"
                
                ticket_text = f"""📨 **تیکت جدید**

👤 **از طرف:** [{username}](tg://user?id={user_id})
🆔 **شناسه کاربر:** `{user_id}`

📝 **پیام:**
{message}

⏰ **زمان ارسال:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"""
                await client.send_message(log_channel_id, f"💬 تیکت جدید: {user_id},\nپیام: {message}")
                keyboard = [
                    [Button.inline("📤 پاسخ", f"reply_{user_id}"), Button.inline("❌ بستن تیکت", f"close_{user_id}")]
                ]
                with open('config.json') as f:
                    config = json.load(f)
                    admin_id = config['ticket_admin_id']
                try:
                    await client.send_message(
                        admin_id,
                        ticket_text,
                        buttons=keyboard,
                        parse_mode='markdown'
                    )
                    await conv.send_message("✅ پیام شما با موفقیت به ادمین ارسال شد. لطفا منتظر پاسخ بمانید.")
                except Exception as e:
                    print(f"Error sending ticket to admin: {e}")
                    await conv.send_message("❌ خطا در ارسال تیکت به ادمین! لطفا بعدا دوباره تلاش کنید.")
                    return
                    
            except asyncio.TimeoutError:
                await conv.send_message("⏰ زمان ارسال پیام به پایان رسید. لطفا دوباره تلاش کنید.")
                return
                
    except Exception as e:
        print(f"Error in ticket handler: {e}")
        await event.reply("❌ خطا در سیستم تیکت! لطفا دوباره تلاش کنید.")

@client.on(events.CallbackQuery(pattern=r"reply_"))
async def reply_handler(event):
    try:
        user_id = event.data.decode().split('_')[1]
        await event.edit("📝 لطفا پاسخ خود را بنویسید:")
        
        async with client.conversation(event.chat_id, timeout=300) as conv:
            try:
                await conv.send_message("📝 لطفا پاسخ خود را بنویسید:")
                
                admin_response = await conv.get_response()
                print(admin_response)
                response_text = f"""📨 **پاسخ ادمین:**

{admin_response.text}

⏰ زمان پاسخ: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"""

                try:
                    await client.send_message(
                        int(user_id), 
                        response_text,
                        parse_mode='markdown'
                    )
                    await event.edit("✅ پاسخ شما با موفقیت ارسال شد.")
                    await client.send_message(log_channel_id, f"💬 پاسخ ادمین به تیکت: {user_id}\n\n{response_text}")

                except Exception as e:
                    print(f"Error sending admin response: {e}")
                    await event.edit("❌ خطا در ارسال پاسخ به کاربر!")
                    
            except asyncio.TimeoutError:
                await event.edit("⏰ زمان پاسخگویی به پایان رسید. لطفا دوباره تلاش کنید.")
                
    except Exception as e:
        print(f"Error in reply handler: {e}")
        await event.edit("❌ خطا در پردازش پاسخ!")

@client.on(events.NewMessage(pattern='/export_data'))
async def export_data_handler(event):
    try:
        # Check if sender is admin
        with open('config.json', 'r') as f:
            config = json.load(f)
            admin_ids = config.get('admin_ids', [])
        if event.sender_id not in admin_ids:
            await event.reply("❌ شما دسترسی به این دستور را ندارید!")
            return

        await event.reply("📊 در حال استخراج اطلاعات...")
        
        db = Database()
        
        # Get data from database
        users_with_wallets = db.get_all_users_with_wallets()
        games = db.get_all_games() 
        bet_games = db.get_all_bet_games()

        # Create Excel workbook
        wb = Workbook()
        
        # Users worksheet
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append([
            "User ID", "Username", "Balance", "Wins", "Losses", "Draws", 
            "Balance FR", "Wallet Address"
        ])
        for user in users_with_wallets:
            # Exclude the wallet key (last element)
            ws_users.append(user[:-1])

        # Games worksheet  
        ws_games = wb.create_sheet("Games")
        ws_games.append([
            "ID", "Game ID", "Board", "Current Player", "Game Over",
            "Player 1 ID", "Player 1 Message ID", "Player 2 ID", "Player 2 Message ID",
            "Is Bet", "Type Game", "Bet Amount", "Winner ID", "Moves",
            "Started At", "Ended At", "Last Move At"
        ])
        for game in games:
            ws_games.append(game)

        # Bet Games worksheet
        ws_bet = wb.create_sheet("Bet Games") 
        ws_bet.append([
            "ID", "Game ID", "Board", "Current Player", "Game Over",
            "Player 1 ID", "Player 1 Message ID", "Player 2 ID", "Player 2 Message ID", 
            "Is Bet", "Type Game", "Bet Amount", "Winner ID", "Moves",
            "Started At", "Ended At", "Last Move At"
        ])
        for game in bet_games:
            ws_bet.append(game)

        # Referral counts worksheet
        ws_ref = wb.create_sheet("Referrals")
        ws_ref.append(["User ID", "Referral Count"])
        with open('referral_counts.json') as f:
            ref_data = json.load(f)
            for user_id, count in ref_data.items():
                ws_ref.append([user_id, count])

        # Save workbook
        filename = f"game_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

        # Send file
        await client.send_file(
            event.chat_id,
            filename,
            caption="📊 فایل اکسل حاوی اطلاعات کامل بازی‌ها، کاربران و کیف پول‌ها"
        )

        # Delete temporary file
        os.remove(filename)

    except Exception as e:
        print(f"Error exporting data: {e}")
        await event.reply("❌ خطا در استخراج اطلاعات!")



async def main():
    try:
        
        await client.start(bot_token=bot_token)
        print("ربات در حال اجراست...")
        await client.run_until_disconnected()
    except Exception as e:
        print(f"خطا: {e}")
    finally:
        await client.disconnect()

if __name__ == "__main__":
    asyncio.run(main())
